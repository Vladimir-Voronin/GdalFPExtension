import datetime
import math
import time
import tracemalloc

from openpyxl import Workbook
from qgis.core import *
from algorithms.GdalUAV.transformation.coordinates.CoordinateTransform import CoordinateTransform
from ModuleInstruments.DebugLog import DebugLog
from algorithms.GdalUAV.processing.FindPathData import FindPathData
from algorithms.AStarMethod import AStarMethod
from algorithms.GdalUAV.processing.calculations.ObjectsCalculations import length_of_path_from_feats_lines
import csv
import openpyxl
import pandas as pd

from algorithms.GdalUAV.exceptions.MethodsException import FailFindPathException, TimeToSucceedException
from pympler import muppy

from algorithms.RandomizedRoadmapMethod import RandomizedRoadmapMethod
from algorithms.SeparationMethod import SeparationMethod

all_objects = muppy.get_objects()


def average(l):
    if len(l) == 0:
        return 0
    return sum(l) / len(l)


class Couple:
    def __init__(self, point_start, point_target):
        self.point_start = point_start
        self.point_target = point_target
        self.base_length = self.__get_length()

    def __get_length(self):
        point_1 = self.point_start.asPoint()
        point_2 = self.point_target.asPoint()
        x_full_difference = point_2.x() - point_1.x()
        y_full_difference = point_2.y() - point_1.y()
        result = math.sqrt(x_full_difference ** 2 + y_full_difference ** 2)
        return (result * result) ** 0.5


class CsvReader:
    @staticmethod
    def get_couple_of_points(file):
        list_of_couples = []
        with open(file, mode='r', newline='') as csv_file:
            csv_reader = csv.reader(csv_file)
            line_count = 0
            for row in csv_reader:
                a = ", ".join(row)
                splited = a.split(';')
                res = []
                for i in splited:
                    res.append(i.split(','))

                for i in range(len(res)):
                    for k in range(len(res[i])):
                        if res[i][k] != "":
                            res[i][k] = float(res[i][k])

                list_of_couples.append([QgsGeometry.fromPointXY(QgsPointXY(res[0][0], res[0][1])),
                                        QgsGeometry.fromPointXY(QgsPointXY(res[1][0], res[1][1]))])
                line_count += 1
            print(f'Processed {line_count} lines.')

        result_list = []
        for i in list_of_couples:
            result_list.append(Couple(i[0], i[1]))

        return result_list


class Test:
    @staticmethod
    def run_test_from_plugin(list_of_pares, method, proj, obstacle_layer, save_doc_path, time_to_stop=200):
        doc_name = "test"
        now = datetime.datetime.now()
        full_name = doc_name + '_' + method.__name__ + str(now.year) + '_' + str(now.month) + '_' + str(
            now.day) + '_' + str(now.hour) + '_' + str(now.minute) + '_' + str(now.second)

        full_name_xlsx = save_doc_path + '/' + full_name + '.xlsx'
        full_name_csv = save_doc_path + '/' + full_name + '.csv'
        wb = Workbook(full_name_xlsx)
        wb.save(full_name_xlsx)

        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(full_name_xlsx)

        list_of_couples = []
        for pare in list_of_pares:
            c = Couple(QgsGeometry.fromPointXY(QgsPointXY(pare.x1, pare.y1)),
                       QgsGeometry.fromPointXY(QgsPointXY(pare.x2, pare.y2)))
            list_of_couples.append(c)

        try:
            # Define variable to read the active sheet:
            worksheet = wookbook.active
            worksheet['A1'] = method.__name__
            worksheet['A2'] = "Points Distance"
            worksheet['B2'] = "Time"
            worksheet['C2'] = "Memory"
            worksheet['D2'] = "Path length"
            worksheet['E2'] = "Numbers of objects"

            worksheet['H1'] = f"{method.__name__} Average"
            worksheet['H2'] = "Points Distance"
            worksheet['I2'] = "Av Time"
            worksheet['J2'] = "Av Memory"
            worksheet['K2'] = "Av Path"
            worksheet['L2'] = "Av Path %"
            worksheet['M2'] = "Av Obstacles"
            worksheet['N2'] = "Succeed precent"

            worksheet['P1'] = f"{method.__name__} Fail"
            worksheet['P2'] = "Points Distance"
            worksheet['Q2'] = "Point1"
            worksheet['R2'] = "Point2"
            worksheet['S2'] = "Type of Error"

            current_row_succeed = 3
            current_row_average = 3
            current_row_fail = 3

            source_list_of_geometry_obstacles = CoordinateTransform.get_list_of_poligons_in_3395(obstacle_layer, proj)

            previous_length = 10000000

            # Statistic for analysis
            time_usage_list = []
            memory_usage_list = []
            length_of_path_list = []
            number_of_obstacles_list = []
            succeed_counter = 0
            fail_counter = 0

            for points in list_of_couples:
                full_time = 0
                full_memory = 0
                result = True
                error = None
                point1 = points.point_start
                point2 = points.point_target
                current_length = points.base_length

                # Average
                if previous_length + 2 < current_length:
                    time_av = average(time_usage_list)
                    memory_av = average(memory_usage_list)
                    length_av = average(length_of_path_list)
                    numb_of_obst_av = average(number_of_obstacles_list)

                    worksheet.cell(current_row_average, 8, value=previous_length)
                    worksheet.cell(current_row_average, 9, value=time_av)
                    worksheet.cell(current_row_average, 10, value=memory_av)
                    worksheet.cell(current_row_average, 11, value=length_av)
                    worksheet.cell(current_row_average, 12, value=f"{length_av / previous_length}")
                    worksheet.cell(current_row_average, 13, value=numb_of_obst_av)
                    worksheet.cell(current_row_average, 14,
                                   value=f"{succeed_counter / (succeed_counter + fail_counter)}")

                    time_usage_list.clear()
                    memory_usage_list.clear()
                    length_of_path_list.clear()
                    number_of_obstacles_list.clear()
                    succeed_counter = 0
                    fail_counter = 0
                    current_row_average += 1

                previous_length = current_length

                find_path_data = FindPathData(proj, point1, point2, obstacle_layer,
                                              save_doc_path,
                                              False,
                                              source_list_of_geometry_obstacles)
                debug_log = DebugLog()
                my_time = time.perf_counter()
                check = method(find_path_data, debug_log)
                try:
                    # starting the monitoring
                    tracemalloc.start()
                    _, start_peak = tracemalloc.get_traced_memory()
                    my_time = time.perf_counter()
                    check.run()
                    _, final_peak = tracemalloc.get_traced_memory()
                    tracemalloc.stop()
                    full_memory = final_peak - start_peak

                except FailFindPathException:
                    error = FailFindPathException.__name__
                    result = False

                except TimeToSucceedException:
                    error = TimeToSucceedException.__name__
                    result = False

                except Exception as e:
                    error = str(e)
                    result = False

                number_of_obstacles = len(check.list_of_obstacles_geometry)
                my_time = time.perf_counter() - my_time
                full_time += my_time

                # Succeed
                if result:
                    succeed_counter += 1
                    length_full = length_of_path_from_feats_lines(check.final_path)

                    time_usage_list.append(full_time)
                    memory_usage_list.append(full_memory)
                    number_of_obstacles_list.append(number_of_obstacles)
                    length_of_path_list.append(length_full)

                    worksheet.cell(current_row_succeed, 1, value=points.base_length)
                    worksheet.cell(current_row_succeed, 2, value=full_time)
                    worksheet.cell(current_row_succeed, 3, value=full_memory)
                    worksheet.cell(current_row_succeed, 4, value=length_full)
                    worksheet.cell(current_row_succeed, 5, value=number_of_obstacles)

                    current_row_succeed += 1
                # Fail
                else:
                    fail_counter += 1
                    worksheet.cell(current_row_fail, 16, value=points.base_length)
                    worksheet.cell(current_row_fail, 17, value=f"{point1.asPoint().x()}, {point1.asPoint().y()}")
                    worksheet.cell(current_row_fail, 18, value=f"{point2.asPoint().x()}, {point2.asPoint().y()}")
                    worksheet.cell(current_row_fail, 19, value=error)
                    current_row_fail += 1

                wookbook.save(full_name_xlsx)
                # if full_time > time_to_stop:
                #     break

            time_av = average(time_usage_list)
            memory_av = average(memory_usage_list)
            length_av = average(length_of_path_list)
            numb_of_obst_av = average(number_of_obstacles_list)

            worksheet.cell(current_row_average, 8, value=previous_length)
            worksheet.cell(current_row_average, 9, value=time_av)
            worksheet.cell(current_row_average, 10, value=memory_av)
            worksheet.cell(current_row_average, 11, value=length_av)
            worksheet.cell(current_row_average, 12, value=f"{length_av / previous_length}")
            worksheet.cell(current_row_average, 13, value=numb_of_obst_av)
            worksheet.cell(current_row_average, 14,
                           value=f"{succeed_counter / (succeed_counter + fail_counter)}")
            wookbook.save(full_name_xlsx)
            read_file = pd.read_excel(f'{full_name_xlsx}')
            read_file.to_csv(f'{full_name_csv}', index=None, header=True)

        finally:
            wookbook.close()


@staticmethod
def run_test_from_code(list_of_couples, methods_list):
    QgsApplication.setPrefixPath(r'C:\OSGEO4~1\apps\qgis', True)
    qgs = QgsApplication([], False)
    qgs.initQgis()

    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(r"C:\Users\Neptune\Desktop\results.xlsx")
    try:
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        current_row = worksheet.max_row + 1
        for method in methods_list:
            point_number = 0
            for points in list_of_couples:
                print(f"start: {points.point_start}")
                print(f"target: {points.point_target}")
                print(f"point_numb: {point_number}")
                point_number += 1
                if point_number < 0:
                    continue
                n = 1
                length_full = 0
                full_time = 0
                number_of_obstacles = 1
                result = True
                full_memory = 0
                area_precent = 0
                for i in range(n):
                    proj = QgsProject.instance()
                    proj.read(r'C:\Users\Neptune\Desktop\Voronin qgis\Voronin qgis.qgs')
                    point1 = points.point_start
                    point2 = points.point_target
                    path = r"C:\Users\Neptune\Desktop\Voronin qgis\shp\Строения.shp"

                    obstacles = QgsVectorLayer(path)
                    source_list_of_geometry_obstacles = CoordinateTransform.get_list_of_poligons_in_3395(obstacles, proj)
                    find_path_data = FindPathData(proj, point1, point2, obstacles,
                                                  r"C:\Users\Neptune\Desktop\Voronin qgis\shp",
                                                  False,
                                                  source_list_of_geometry_obstacles)
                    debug_log = DebugLog()
                    my_time = time.perf_counter()
                    check = method(find_path_data, debug_log)
                    try:
                        # starting the monitoring
                        tracemalloc.start()
                        _, start_peak = tracemalloc.get_traced_memory()
                        check.run()
                        _, final_peak = tracemalloc.get_traced_memory()
                        tracemalloc.stop()
                        full_memory = final_peak - start_peak
                        # area_precent = check.get_area_precents()

                    except QgsException:
                        result = False

                    number_of_obstacles = len(check.list_of_obstacles_geometry)
                    my_time = time.perf_counter() - my_time
                    full_time += my_time
                    if not check.final_path:
                        result = False
                    else:
                        length_full += length_of_path_from_feats_lines(check.final_path)

                if not result:
                    length_full = 0
                    full_time = 0
                    number_of_obstacles = 0
                full_time /= n
                full_memory /= n
                number_of_obstacles /= n
                print(method.__name__)
                print(full_time)
                print(f"{full_memory} b")
                print(f"n obst: {number_of_obstacles}")
                length_full /= n
                print(f"Length: {length_full}")
                worksheet.cell(current_row, 1, value=points.base_length)
                worksheet.cell(current_row, 2, value=method.__name__)
                worksheet.cell(current_row, 3, value=full_time)
                worksheet.cell(current_row, 4, value=length_full)
                worksheet.cell(current_row, 5, value=n)
                worksheet.cell(current_row, 6, value=full_memory)
                worksheet.cell(current_row, 7, value=number_of_obstacles)
                # worksheet.cell(current_row, 8, value=area_precent)

                current_row += 1
                wookbook.save(r"C:\Users\Neptune\Desktop\results.xlsx")
                if full_time > 90:
                    break
    finally:
        wookbook.close()


if __name__ == '__main__':
    Test.run_test_from_plugin([], AStarMethod, None, None, r"C:\Users\Neptune\Desktop")
